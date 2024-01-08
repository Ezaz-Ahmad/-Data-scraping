import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def find_us_subsidiaries(link):
    states = ["Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware",
              "Florida", "Georgia",
              "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland",
              "Massachusetts",
              "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire",
              "New Jersey",
              "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania",
              "Rhode Island",
              "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
              "West Virginia",
              "Wisconsin", "Wyoming"]

    state_counts = {state: 0 for state in states}

    headers = {'User-Agent': 'Mozilla/5.0 ...'}
    session = requests.Session()
    session.headers.update(headers)

    try:
        response = session.get(link)
        if response.status_code != 200:
            return f"Failed to retrieve the webpage, Status code: {response.status_code}"

        soup = BeautifulSoup(response.content, 'html.parser')

        # Check for the format of the table on the webpage
        if "Arthur J. Gallagher" in response.text:  # Checking for a unique identifier from the new link
            # New link parsing logic
            for row in soup.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) >= 3:  # Ensure there are at least 3 cells in the row
                    location = cells[2].get_text(strip=True)
                    for state in states:
                        if state.lower() in location.lower():
                            state_counts[state] += 1
                            break
        else:
            # Original link parsing logic
            for table in soup.find_all('table'):
                for row in table.find_all('tr'):
                    cells = row.find_all('td')
                    if len(cells) > 1:
                        location = cells[1].get_text(strip=True)
                        for state in states:
                            if state.lower() in location.lower():
                                state_counts[state] += 1
                                break

        return state_counts

    except requests.exceptions.RequestException as e:
        return f"An error occurred: {e}"

def append_to_excel(df, path, sheet_name='Sheet1'):
    if os.path.exists(path):
        # Load the workbook
        book = openpyxl.load_workbook(path)

        # Check if the sheet exists, if not create it
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)

        # Find the starting row
        startrow = sheet.max_row + 1

        # Append the data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), startrow):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

                # Highlight the first row in yellow
                if r_idx == startrow:
                    sheet.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Save the workbook
        book.save(path)
    else:
        # Write the DataFrame to a new Excel file and apply highlighting
        df.to_excel(path, index=False)
        book = openpyxl.load_workbook(path)
        sheet = book[sheet_name]
        for cell in sheet['A1:B1'][0]:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        book.save(path)
# working with the user provided link
link = input("Enter the URL: ")
us_subsidiaries = find_us_subsidiaries(link)

df = pd.DataFrame(list(us_subsidiaries.items()), columns=['State', 'Count'])

downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
excel_file_path = os.path.join(downloads_path, 'Data.xlsx')

append_to_excel(df, excel_file_path)

print(f"Data appended to {excel_file_path}")

